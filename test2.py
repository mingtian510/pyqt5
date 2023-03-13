import sys
from PyQt5.QtWidgets import QApplication, QLabel, QTableView, QVBoxLayout, QLineEdit, QHBoxLayout, QWidget, \
    QMainWindow, QAction, QDialog, QDialogButtonBox, QPlainTextEdit, QTextEdit, QPushButton
from PyQt5.QtGui import QIcon, QFont, QFontMetrics
from PyQt5.QtCore import QSettings, Qt, QSortFilterProxyModel
from PyQt5.QtGui import QStandardItemModel, QStandardItem
from logzero import logger
import xlrd
from PyQt5.QtGui import QPainter
from config_contrast_public import txt_to_xlsx, get_config_field, check_config_result
from openpyxl import load_workbook
import re


class ConfigBaseData:
    client_path = ''
    client_id_col = 0
    client_field_col = 0
    server_path = ''
    server_pos_start_str = ''
    server_pos_end_str = ''
    server_id_re = ''
    server_field_re = ''

# 要重写QSortFilterProxyModel类
class SortFilterProxyModel(QSortFilterProxyModel):

    def __init__(self, *args, **kwargs):
        QSortFilterProxyModel.__init__(self, *args, **kwargs)
        self.filter_column = 0  # 设置默认过滤列为第一列

    def setFilterColumn(self, column):
        self.filter_column = column

    def filterAcceptsRow(self, source_row, source_parent):
        index = self.sourceModel().index(source_row, self.filter_column, source_parent)
        data = self.sourceModel().data(index)
        if self.filterRegExp().indexIn(data) != -1:
            return True
        else:
            return False


class LineNumberArea(QTextEdit):
    def __init__(self, editor):
        super().__init__(editor)
        self.editor = editor
        self.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.setStyleSheet("background-color: #e0e0e0;")

    def update_width(self):
        width = self.fontMetrics().width(str(self.editor.blockCount())) + 10
        self.setFixedWidth(width)

    def update(self, rect, dy):
        if dy:
            self.scroll(0, dy)
        else:
            self.updateGeometry()
        if rect.contains(self.editor.viewport().rect()):
            self.editor.setViewportMargins(self.width(), 0, 0, 0)

        super().update(rect)

    def paintEvent(self, event):
        self.editor.setViewportMargins(self.width(), 0, 0, 0)
        block = self.editor.firstVisibleBlock()
        block_number = block.blockNumber()
        painter = QPainter(self.viewport())
        painter.fillRect(event.rect(), Qt.lightGray)
        font = painter.font()
        font.setPointSize(10)
        painter.setFont(font)
        current_height = self.fontMetrics().height()
        top = self.editor.blockBoundingGeometry(block).translated(self.editor.contentOffset()).top()
        bottom = top + self.editor.blockBoundingRect(block).height()
        while block.isValid() and top <= event.rect().bottom():
            if block.isVisible() and bottom >= event.rect().top():
                number = str(block_number + 1)
                painter.drawText(0, int(top), int(self.width()) - 5, current_height, Qt.AlignRight, number)
            block = block.next()
            top = bottom
            bottom = top + self.editor.blockBoundingRect(block).height()
            block_number += 1
        painter.end()
        super().paintEvent(event)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.config_name = None
        self.settings = QSettings('Youkia', 'Config_Filter')
        logger.info(self.settings.value('config_dict'))
        if self.settings.value('config_dict'):
            self.config_dict = self.settings.value('config_dict')
        else:
            self.config_dict = {}
        # Create main widget
        central_widget = QWidget(self)
        self.setCentralWidget(central_widget)

        # 前端配置数据
        self.client_id_label = QLabel("前端ID标题：")
        self.client_id_text = QLineEdit()

        self.client_field_label = QLabel("数据列标题")
        self.client_field_text = QLineEdit()
        # 如果有配置名字，初始化保存好的本地数据
        # if self.config_name:
        #     self.client_id_text.setText(self.config_dict[self.config_name]['client_id_col'])
        #     self.client_field_text.setText(self.config_dict[self.config_name]['client_field_col'])
        self.client_field_text.setPlaceholderText("请输入前端配置表所需数据列标题")
        self.client_field_text.setAlignment(Qt.AlignCenter)

        # 后端配置数据
        self.server_pos_start_label = QLabel("后端开始行文字")
        self.server_pos_start_text = QLineEdit()

        self.server_pos_end_label = QLabel("后端结束行文字")
        self.server_pos_end_text = QLineEdit()

        self.server_id_re_label = QLabel("后端ID数据正则表达式")
        self.server_id_re_text = QLineEdit()

        self.server_field_re_label = QLabel("后端字段数据正则表达式")
        self.server_field_re_text = QLineEdit()


        config_form_layout = QHBoxLayout()
        # 前端
        config_form_layout.addWidget(self.client_id_label)
        config_form_layout.addWidget(self.client_id_text)
        config_form_layout.addWidget(self.client_field_label)
        config_form_layout.addWidget(self.client_field_text)

        # 后端
        config_form_layout.addWidget(self.server_pos_start_label)
        config_form_layout.addWidget(self.server_pos_start_text)
        config_form_layout.addWidget(self.server_pos_end_label)
        config_form_layout.addWidget(self.server_pos_end_text)
        config_form_layout.addWidget(self.server_id_re_label)
        config_form_layout.addWidget(self.server_id_re_text)
        config_form_layout.addWidget(self.server_field_re_label)
        config_form_layout.addWidget(self.server_field_re_text)


        # 保存按钮
        self.submit_button = QPushButton('保存配置设置', self)
        self.submit_button.clicked.connect(self.submit_config_setting)

        config_form_layout.addWidget(self.submit_button)

        # Create table view
        self.table_view_client = QTableView()
        self.table_view_server = QPlainTextEdit()
        self.table_view_server.setReadOnly(True)


        table_layout = QHBoxLayout()
        table_layout.addWidget(self.table_view_client)
        table_layout.addWidget(self.table_view_server)

        # 配置对比按钮
        self.check_button = QPushButton('前后配置比对', self)
        self.check_button.clicked.connect(self.check_config)

        check_button_layout = QHBoxLayout()
        check_button_layout.addWidget(self.check_button)

        # 结果文本框
        self.result_text = QTextEdit()
        self.result_text.setReadOnly(True)

        result_layout = QHBoxLayout()
        result_layout.addWidget(self.result_text)


        # 主视图
        main_layout = QVBoxLayout(central_widget)
        main_layout.addLayout(config_form_layout)
        main_layout.addLayout(table_layout)
        main_layout.addLayout(check_button_layout)
        main_layout.addLayout(result_layout)

        self.statusBar()

        # # 创建打开文件的动作
        # openFile = QAction('打开', self)
        # openFile.setShortcut('Ctrl+O')

        # 创建设置配置文件的动作
        set_file_path_action = QAction('设置配置表路径', self)
        set_file_path_action.setShortcut('Ctrl+O')
        set_file_path_action.triggered.connect(self.set_config_file_path)

        # 创建菜单栏
        menubar = self.menuBar()
        fileMenu = menubar.addMenu('文件')
        fileMenu.addAction(set_file_path_action)

        localconfigMenu = menubar.addMenu('配置表选择')
        for name in self.config_dict:
            action = QAction(name, self)
            action.triggered.connect(self.select_config_name)
            localconfigMenu.addAction(action)


        if self.config_name:
            self.init_local_data()

        # Set window properties
        self.setGeometry(100, 100, 800, 600)
        self.setWindowIcon(QIcon('demo2.ico'))
        self.setWindowTitle("表格筛选")

    def init_table(self, file_path, file_type):
        if file_type == 'client':
            # 设置表格视图属性
            # self.table_view_client.setSortingEnabled(True)
            # self.table_view_client.setEditTriggers(QTableView.NoEditTriggers)
            # self.table_view_client.setSelectionBehavior(QTableView.SelectRows)
            # self.table_view_client.setSelectionMode(QTableView.SingleSelection)
            if file_path.endswith('.txt'):
                with open(r'' + file_path, 'r', encoding='UTF-8') as f:
                    try:
                        content = f.read()
                    except Exception as e:
                        self.statusBar().showMessage("文件读取失败：" + str(e))
                    _list = content.split('\n')
                    title_row = _list[1].split("\t")
                    # title_row = [i for i in title_row if i.strip() != ""]
                    self.model = QStandardItemModel(len(_list) - 2, len(title_row))
                    self.model.setHorizontalHeaderLabels(title_row)
                    self.table_view_client.setModel(self.model)

                    # 在表格控件中显示列内容
                    try:
                        for m, val in enumerate(_list[3:]):
                            if val:
                                # val 为每行的数据
                                cell_val_list = val.split("\t")
                                for i, cell in enumerate(cell_val_list):
                                    item = QStandardItem(str(cell))
                                    self.model.setItem(m, i, item)
                            else:
                                continue
                    except Exception as e:
                        self.statusBar().showMessage("初始化表格数据失败：" + str(e))
            elif file_path.endswith('.xlsx') or file_path.endswith('.xls') or file_path.endswith('.xlsm'):
                workbook = xlrd.open_workbook(file_path)
                work_sheet = workbook.sheet_by_index(0)
                nrows = work_sheet.nrows
                ncols = work_sheet.ncols
                title_row = work_sheet.row_values(1)
                self.model = QStandardItemModel(nrows - 2, ncols)
                self.model.setHorizontalHeaderLabels(title_row)
                self.table_view_client.setModel(self.model)

                # 在表格控件中显示列内容
                try:
                    for n in range(2, nrows):
                        cell_val_list = work_sheet.row_values(n)
                        for i, cell in enumerate(cell_val_list):
                            c_type = work_sheet.cell(n, i).ctype
                            if c_type == 2 and cell % 1 == 0.0:  # 处理整数变为浮点数的情况， ctype为2且为浮点
                                cell = int(cell)
                            item = QStandardItem(str(cell))
                            self.model.setItem(n, i, item)
                except Exception as e:
                    self.statusBar().showMessage("初始化表格数据失败：" + str(e))

        elif file_type == 'server':
            try:
                if file_path.endswith('.cfg'):
                    with open(r'' + file_path, 'r', encoding='UTF-8') as f:
                        try:
                            content = f.read()
                        except Exception as e:
                            self.statusBar().showMessage("文件读取失败：" + str(e))
                        # _list = content.split('\n')
                        self.table_view_server.setPlainText(content)
            except Exception as e:
                self.statusBar().showMessage("初始化后台表数据失败：" + str(e))

    def set_config_file_path(self):
        # 创建一个对话框
        dialog = QDialog(self)
        dialog.setWindowTitle("设置配置表路径")

        # 创建一个垂直布局
        layout = QVBoxLayout(dialog)

        # 创建配置表名字标签和文本框
        config_name_label = QLabel("自定义配置表名：", dialog)
        layout.addWidget(config_name_label)
        config_name_text_Edit = QLineEdit(dialog)
        layout.addWidget(config_name_text_Edit)

        # 创建前端路径标签和文本框
        client_label = QLabel("前端文件路径：", dialog)
        layout.addWidget(client_label)
        client_text_Edit = QLineEdit(dialog)
        layout.addWidget(client_text_Edit)

        # 创建后端路径标签和文本框
        server_label = QLabel("后端文件路径：", dialog)
        layout.addWidget(server_label)
        server_text_Edit = QLineEdit(dialog)
        layout.addWidget(server_text_Edit)

        # 创建一个按钮盒子
        buttonBox = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, dialog)
        buttonBox.accepted.connect(dialog.accept)
        buttonBox.rejected.connect(dialog.reject)
        layout.addWidget(buttonBox)

        # 显示对话框
        if dialog.exec_() == QDialog.Accepted:
            try:
                client_path = client_text_Edit.text()
                server_path = server_text_Edit.text()
                logger.info(self.config_dict)
                # 初始化基础数据
                client_field_col = 0
                client_id_col = 0
                server_pos_start_str = ''
                server_pos_end_str = ''
                server_id_re = ''
                server_field_re = ''
                config_data = dict(
                    client_path=client_path,
                    client_id_col=client_id_col,
                    client_field_col=client_field_col,
                    server_path=server_path,
                    server_pos_start_str=server_pos_start_str,
                    server_pos_end_str=server_pos_end_str,
                    server_id_re=server_id_re,
                    server_field_re=server_field_re
                )
                self.config_dict[config_name_text_Edit.text()] = config_data
                self.config_name = config_name_text_Edit.text()
                self.settings.setValue('config_dict', self.config_dict)
                logger.info(self.config_dict)
                self.select_config_name()
            except Exception as e:
                self.statusBar().showMessage(str(e))

    def select_config_name(self):
        if not self.config_name:
            self.config_name = self.sender().text()
        logger.info(self.config_name)
        client_config_path = self.config_dict[self.config_name]['client_path']
        if client_config_path:
            self.init_table(client_config_path, 'client')
        server_config_path = self.config_dict[self.config_name]['server_path']
        if server_config_path:
            self.init_table(server_config_path, 'server')

    def submit_config_setting(self):
        if self.config_name:
            try:
                headr = self.table_view_client.horizontalHeader()
                column_count = self.model.columnCount()
                client_id_col = -1
                client_field_col = -1
                for i in range(column_count):
                    header_data = self.model.headerData(i, Qt.Horizontal)
                    if header_data == self.client_id_text.text():
                        client_id_col = headr.visualIndex(i) + 1
                    if header_data == self.client_field_text.text():
                        client_field_col = headr.visualIndex(i) + 1
                _dict = dict(
                    client_id_col=client_id_col,
                    client_field_col=client_field_col,
                    server_pos_start_str=self.server_pos_start_text.text(),
                    server_pos_end_str=self.server_pos_end_text.text(),
                    server_id_re=self.server_id_re_text.text(),
                    server_field_re=self.server_field_re_text.text()
                )
                self.config_dict[self.config_name].update(_dict)
                logger.info(self.config_dict[self.config_name])
                self.settings.setValue('config_dict', self.config_dict)
            except Exception as e:
                logger.info(e)
                self.statusBar().showMessage("报错设置失败：" + str(e))

        else:
            self.statusBar().showMessage('未选中对应配置表')

    def check_config(self):
        config_data = self.config_dict[self.config_name]
        client_path = self.config_dict[self.config_name]['client_path']
        server_path = self.config_dict[self.config_name]['server_path']
        client_file = self.config_name + '_client.xlsx'
        txt_to_xlsx(client_path, client_file)
        client_workbook = load_workbook(filename=client_file, data_only=True)

        client_sheet = client_workbook['sheet1']

        # # 获取字段名字典
        # field_dict = get_config_field(qian_sheet)
        #
        # # 获取堆叠数量列数
        # qian_pilenum_col = field_dict['堆叠数量']

        # # 获取出售价格列数
        # qian_value_col = field_dict['出售价格']

        client_result_dict = {}
        # 获取道具sid
        for cols in client_sheet.iter_cols(min_row=50, values_only=False):  # 去掉前台配置表中多余的废弃道具，所以从50行开始取数据
            logger.info(cols[0].column)
            logger.info(config_data['client_id_col'])
            # print(cols)
            if cols[0].column == config_data['client_id_col']:
                for cell in cols:
                    logger.info(cell.value)
                    client_value = client_sheet.cell(row=cell.row,
                                                   column=(cell.column + config_data['client_field_col'] - config_data['client_id_col'])).value
                    if client_value is None:
                        client_value = '0'
                    client_result_dict[cell.value] = '{}'.format(client_value)
        logger.info(client_result_dict)

        # 处理后台数据，这里直接把后台配置表需要的数据截取出来，也可以把后台配置表处理为xlsx文件再处理
        # f = open('./hou/goods.cfg')
        with open(r'' + server_path, 'r', encoding='UTF-8') as f:
            server_content = f.read()
        # print(re.search('{.+?keypos', data).group())
        pos_start = server_content.find(config_data['server_pos_start_str'])
        pos_end = server_content.find(config_data['server_pos_end_str'])
        data_new = server_content[pos_start + len(config_data['server_pos_start_str']):pos_end]
        data_new = data_new.replace(' ', '').split('\n')[2:-2]
        server_result_dict = {}
        for _data in data_new:
            # 取道具sid数据
            server_sid = re.findall(config_data['server_id_re'], _data)
            # 取道具堆叠上限数据
            server_value = re.findall(config_data['server_field_re'], _data)
            if server_value == [] and len(server_value) == 0:
                server_value = ['0']
            server_result_dict[server_sid[0]] = "{}".format(server_value[0])
        logger.info(server_result_dict)
        # 前后数据对比结果
        result = check_config_result(client_result_dict, server_result_dict, self.config_name)
        result = result + '\n' + str(client_result_dict) + '\n' + str(server_result_dict)
        self.result_text.setText(result)




if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
