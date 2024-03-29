
from openpyxl import Workbook
from logzero import logger

def check_config_result(client_data, server_data, result=""):
    res = result + '\n'
    if not len(client_data):
        res += "前端未找到对应字段数据，请检查配置内容 \n"
        return res
    if not len(server_data):
        res += "后端未找到对应字段数据，请检查配置内容 \n"
        return res
    if len(client_data) == len(server_data):
        for key1 in client_data:
            value1 = client_data.get(key1)
            value2 = server_data.get(key1)
            if value2:
                if value1 != value2:
                    res += '不一致的sid：{}，前台配置：{}，后台配置：{} \n'.format(key1, value1, value2)
            else:
                res += 'sid不一样，前端sid：{} \n'.format(key1)

    else:
        res += "前后配置长度不一致， 前端有{}条数据， 后端有{}条数据 \n".format(len(client_data), len(server_data))
        # 用长度长的作为基准
        if len(client_data) > len(server_data):
            for key1 in client_data:
                value1 = client_data.get(key1)
                value2 = server_data.get(key1)
                if value2:
                    if value1 != value2:
                        res += '不一致的sid：{}，前台配置：{}，后台配置：{} \n'.format(key1, value1, value2)
        else:
            for key1 in server_data:
                value1 = server_data.get(key1)
                value2 = client_data.get(key1)
                if value2:
                    if value1 != value2:
                        res += '不一致的sid：{}，前台配置：{}，后台配置：{} \n'.format(key1, value2, value1)

    return res


def txt_to_xlsx(file_txt_path, file_xlsx_name):
    try:
        # file_txt_path1 = file_txt_path +
        file_txt = open(file_txt_path, 'r', encoding='utf-8')
        xlsx_wb = Workbook()
        sheet = xlsx_wb.create_sheet('sheet1')
        x = 1
        while True:
            # 按行循环，读取文本文件
            line = file_txt.readline()
            if not line:
                break  # 如果没有内容，则退出循环
            for i in range(0, len(line.split('\t'))):
                item = line.split('\t')[i]
                sheet.cell(row=x, column=i + 1, value=item)  # x单元格经度，i单元格纬度
            x += 1  # excel另起一行
        file_txt.close()
        xlsx_wb.save(file_xlsx_name)
        # print("保存成功！")
    except:
        raise


def get_config_field(sheet, row=2):
    """
    获取某个表第row行的字段名
    :param sheet:
    :param row:取第row行的数据
    :return:返回一个字典，key为字段名，value为列数
    """
    field = {}
    max_column = sheet.max_column
    for row in sheet.iter_rows(min_col=1, max_col=max_column, min_row=row, max_row=row, values_only=False):
        for cell in row:
            field[cell.value] = cell.column
    return field
